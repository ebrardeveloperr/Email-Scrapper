import requests
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
import openpyxl
from bs4 import BeautifulSoup
import os
import subprocess
import sys
import string
import time
from multiprocessing import Pool, cpu_count, freeze_support
from langdetect import detect, DetectorFactory
from urllib.parse import urljoin, urlparse
DetectorFactory.seed = 0

# Optional: LLM integration via environment variables
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

def detect_website_language(html_content):
    """Detects the primary language of the webpage."""
    soup = BeautifulSoup(html_content, 'html.parser')
    text = " ".join(soup.stripped_strings)  # Extract visible text
    try:
        lang = detect(text)
        return lang
    except:
        return None

def clean_email(email):
    """Removes everything after '?' (query parameters) from an email."""
    return email.split('?')[0].strip()

def normalize_obfuscated_email(email_text):
    """Converts obfuscated email formats to standard format."""
    if not email_text:
        return ""
    
    # Remove extra whitespace
    email_text = re.sub(r'\s+', ' ', email_text.strip())
    
    # Handle {at} and {dot} patterns
    email_text = re.sub(r'\s*\{at\}\s*', '@', email_text, flags=re.IGNORECASE)
    email_text = re.sub(r'\s*\{dot\}\s*', '.', email_text, flags=re.IGNORECASE)
    
    # Handle [at] and [dot] patterns
    email_text = re.sub(r'\s*\[at\]\s*', '@', email_text, flags=re.IGNORECASE)
    email_text = re.sub(r'\s*\[dot\]\s*', '.', email_text, flags=re.IGNORECASE)
    
    # Handle (at) and (dot) patterns
    email_text = re.sub(r'\s*\(\s*at\s*\)\s*', '@', email_text, flags=re.IGNORECASE)
    email_text = re.sub(r'\s*\(\s*dot\s*\)\s*', '.', email_text, flags=re.IGNORECASE)
    
    # Handle "at" and "dot" patterns
    email_text = re.sub(r'\s+at\s+', '@', email_text, flags=re.IGNORECASE)
    email_text = re.sub(r'\s+dot\s+', '.', email_text, flags=re.IGNORECASE)
    
    # Clean up any remaining extra spaces around @ and .
    email_text = re.sub(r'\s*@\s*', '@', email_text)
    email_text = re.sub(r'\s*\.\s*', '.', email_text)
    
    return email_text.strip()

def clean_phone(raw: str) -> str:
    """Normalize a phone string: keep leading '+' and digits; strip separators.
    Returns empty string if not a plausible length (7-15 digits).
    """
    if not raw:
        return ""
    s = raw.strip()
    if s.lower().startswith('tel:'):
        s = s[4:]
    kept = []
    for ch in s:
        if ch.isdigit() or (ch == '+' and not kept):
            kept.append(ch)
    s = ''.join(kept)
    digits_only = ''.join(ch for ch in s if ch.isdigit())
    if 7 <= len(digits_only) <= 15:
        return s
    return ""

def _log(message: str):
    try:
        log_text.configure(state="normal")
        log_text.insert("end", f"{time.strftime('%H:%M:%S')} - {message}\n")
        log_text.configure(state="disabled")
        log_text.see("end")
    except Exception:
        # Logging not yet available
        pass

def _set_running_state(is_running: bool):
    # Disable/enable inputs during run
    state = "disabled" if is_running else "normal"
    try:
        choose_btn.configure(state=state)
        start_btn.configure(state=state if not is_running else "disabled")
        column_menu.configure(state=state)
        start_entry.configure(state=state)
        end_entry.configure(state=state)
        process_menu.configure(state=state)
        
        # Update status indicator
        # if is_running:
        #     status_indicator.configure(text="● Running", foreground="#ff8c00")
        # else:
        #     status_indicator.configure(text="● Ready", foreground="#107c10")
    except Exception:
        pass

def _auto_open_result(path: str):
    try:
        if os.name == 'nt':
            os.startfile(path)
        elif sys.platform == 'darwin':
            subprocess.run(['open', path], check=False)
        else:
            subprocess.run(['xdg-open', path], check=False)
        _log(f"Opened: {path}")
    except Exception as e:
        _log(f"Could not open file automatically: {e}")

def _on_close():
    try:
        if last_save_path and os.path.exists(last_save_path):
            if os.name == 'nt':
                # Open Explorer selecting the saved file
                subprocess.run(['explorer', f'/select,{last_save_path}'], check=False)
            elif sys.platform == 'darwin':
                subprocess.run(['open', '-R', last_save_path], check=False)
            else:
                subprocess.run(['xdg-open', os.path.dirname(last_save_path)], check=False)
    except Exception:
        pass
    finally:
        try:
            root.destroy()
        except Exception:
            pass

def _llm_suggest_links_from_home(html_text: str, base_url: str, provider: str) -> list:
    """Use an LLM to suggest likely contact/about/imprint URLs on the same domain.

    provider: 'none' | 'openai' | 'gemini'
    Returns absolute URLs (may be empty on error or if keys missing).
    """
    try:
        provider = (provider or "none").lower()
        if provider not in ("openai", "gemini"):
            return []
        # Resolve keys: prefer UI-provided values, fallback to env
        ui_openai = ""
        ui_gemini = ""
        try:
            ui_openai = (globals().get('openai_key_var').get() or "").strip()
        except Exception:
            pass
        try:
            ui_gemini = (globals().get('gemini_key_var').get() or "").strip()
        except Exception:
            pass
        effective_openai = ui_openai or OPENAI_API_KEY
        effective_gemini = ui_gemini or GEMINI_API_KEY
        if provider == "openai" and not effective_openai:
            return []
        if provider == "gemini" and not effective_gemini:
            return []
        prompt = (
            "You are given HTML of a website's home page. Return a concise JSON array of up to 6 \n"
            "absolute URLs on the same site most likely to contain contact information (e.g., \n"
            "Contact, About, Impressum, Support, Team, Company, Privacy pages). Return only the JSON array.\n"
        )
        # Small HTML slice to keep payload short
        snippet = html_text
        if len(snippet) > 6000:
            snippet = snippet[:6000]
        if provider == "openai":
            # Call OpenAI Chat Completions
            headers = {
                "Authorization": f"Bearer {effective_openai}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": "gpt-4o-mini",
                "messages": [
                    {"role": "system", "content": "You output only JSON arrays of URLs, no prose."},
                    {"role": "user", "content": f"Base URL: {base_url}\n\n{prompt}\n\nHTML:\n{snippet}"},
                ],
                "temperature": 0.2,
            }
            resp = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload, timeout=15)
            if resp.status_code != 200:
                return []
            data = resp.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "[]")
        else:
            # Gemini generateContent
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={effective_gemini}"
            payload = {
                "contents": [
                    {
                        "parts": [
                            {"text": "You output only JSON arrays of URLs, no prose."},
                            {"text": f"Base URL: {base_url}\n\n{prompt}\n\nHTML:\n{snippet}"},
                        ]
                    }
                ]
            }
            resp = requests.post(url, json=payload, timeout=15)
            if resp.status_code != 200:
                return []
            data = resp.json()
            content = (
                data.get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [{}])[0]
                .get("text", "[]")
            )
        # Parse JSON array safely
        import json
        urls = json.loads(content)
        abs_urls = []
        base_host = urlparse(base_url).netloc
        for u in urls:
            try:
                abs_u = urljoin(base_url, str(u))
                parsed = urlparse(abs_u)
                if parsed.scheme in ("http", "https") and parsed.netloc.endswith(base_host):
                    abs_urls.append(abs_u)
            except Exception:
                continue
        return abs_urls[:6]
    except Exception:
        return []


def scrape_email(base_url, scrape_options=None):
    """Fetches the HTML of the URL and extracts emails, Facebook profiles, and phone numbers.

    Enhancement: Also follows all on-page links that look like contact pages
    (by link text or href keywords) and scans those as well.
    
    scrape_options: dict with keys 'emails', 'facebook', 'phone', 'count', 'all'
    """
    if scrape_options is None:
        scrape_options = {'emails': True, 'facebook': False, 'phone': False, 'count': True, 'all': True}
    # Enhanced email patterns to handle obfuscated formats
    email_patterns = [
        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b',  # Standard email
        r'\b[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\.\s*[A-Z|a-z]{2,7}\b',  # With spaces
        r'\b[A-Za-z0-9._%+-]+\s*\{at\}\s*[A-Za-z0-9.-]+\s*\{dot\}\s*[A-Z|a-z]{2,7}\b',  # {at} and {dot}
        r'\b[A-Za-z0-9._%+-]+\s*\[at\]\s*[A-Za-z0-9.-]+\s*\[dot\]\s*[A-Z|a-z]{2,7}\b',  # [at] and [dot]
        r'\b[A-Za-z0-9._%+-]+\s*\(\s*at\s*\)\s*[A-Za-z0-9.-]+\s*\(\s*dot\s*\)\s*[A-Z|a-z]{2,7}\b',  # (at) and (dot)
        r'\b[A-Za-z0-9._%+-]+\s*at\s*[A-Za-z0-9.-]+\s*dot\s*[A-Z|a-z]{2,7}\b',  # at and dot
        r'\b[A-Za-z0-9._%+-]+\s*@\s*\{[A-Za-z0-9.-]+\}\s*\.\s*[A-Z|a-z]{2,7}\b',  # @{domain}
        r'\b[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\{dot\}\s*[A-Z|a-z]{2,7}\b',  # @domain{dot}
    ]
    
    # Facebook pattern
    facebook_pattern = r'(?:https?://)?(?:www\.)?facebook\.com/[A-Za-z0-9._-]+'
    
    # Phone pattern
    phone_pattern = r'(?:\+?\d[\d\s().-]{6,}\d)'

    # Seed the queue with base and common contact-like paths
    candidate_urls = [
        base_url,
        f"{base_url}/kontakt",
        f"{base_url}/impressum",
        f"{base_url}/imprint",
        f"{base_url}/impressum-contact",
        f"{base_url}/kontakt-impressum",
        f"{base_url}/contactus",
        f"{base_url}/about",
        f"{base_url}/contact",
        f"{base_url}/cv",
    ]

    # Normalize and dedupe, keep only same-origin http(s)
    def normalize_same_origin(url: str):
        try:
            parsed = urlparse(url)
            return parsed.scheme in ("http", "https")
        except Exception:
            return False

    visited = set()
    queue = []
    # Preload queue with normalized unique candidates
    for u in candidate_urls:
        abs_u = u
        if normalize_same_origin(abs_u) and abs_u not in visited:
            visited.add(abs_u)
            queue.append(abs_u)

    country = ""
    found_emails = set()
    found_facebook_profiles = set()
    first_phone_global = None

    # Contact-related keywords to search in link text or href
    contact_keywords = [
        "contact", "kontakt", "impressum", "imprint", "support", "customer",
        "kundendienst", "about", "unternehmen", "company", "team", "cv",
    ]
    facebook_url = None

    import concurrent.futures
    from threading import current_thread
    # Helper to fetch and parse a single URL for email/phone
    def fetch_and_parse(url):
        if hasattr(globals(), 'stop_event') and globals()['stop_event'].is_set():
            return None
        try:
            response = session_get(url, timeout=10)
            if response.status_code != 200:
                return None
            soup = BeautifulSoup(response.text, 'html.parser')
            # Find Facebook links on this page
            nonlocal facebook_url
            page_facebook_profiles = set()
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                if href.startswith("http") and "facebook.com" in href:
                    if facebook_url is None:
                        facebook_url = href
                    page_facebook_profiles.add(href)

            # Always collect links with 'cv' or 'contact' in URL or link text
            cv_contact_links = set()
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                text_lower = (a.get_text(strip=True) or "").lower()
                href_lower = href.lower()
                if ("cv" in href_lower or "cv" in text_lower or "contact" in href_lower or "contact" in text_lower):
                    abs_url = urljoin(url, href)
                    try:
                        base_host = urlparse(base_url).netloc
                        abs_parsed = urlparse(abs_url)
                        if abs_parsed.scheme in ("http", "https") and abs_parsed.netloc.endswith(base_host):
                            if abs_url not in visited:
                                cv_contact_links.add(abs_url)
                    except Exception:
                        continue
            # Detect language/country from this page (first successful page wins)
            nonlocal country
            if country == "":
                detected_lang = detect_website_language(response.text)
                country = {
                    'af': 'South Africa', 'sq': 'Albania', 'am': 'Ethiopia', 'ar': 'Saudi Arabia',
                    'hy': 'Armenia', 'az': 'Azerbaijan', 'eu': 'Spain', 'bn': 'Bangladesh',
                    'bs': 'Bosnia and Herzegovina', 'bg': 'Bulgaria', 'ca': 'Spain', 'zh-cn': 'China',
                    'zh-tw': 'Taiwan', 'hr': 'Croatia', 'cs': 'Czech Republic', 'da': 'Denmark',
                    'nl': 'Netherlands', 'en': 'United States', 'et': 'Estonia', 'fi': 'Finland',
                    'fr': 'France', 'ka': 'Georgia', 'de': 'Germany', 'el': 'Greece', 'gu': 'India',
                    'he': 'Israel', 'hi': 'India', 'hu': 'Hungary', 'is': 'Iceland', 'id': 'Indonesia',
                    'it': 'Italy', 'ja': 'Japan', 'kn': 'India', 'kk': 'Kazakhstan', 'km': 'Cambodia',
                    'ko': 'South Korea', 'ky': 'Kyrgyzstan', 'lo': 'Laos', 'lv': 'Latvia', 'lt': 'Lithuania',
                    'mk': 'North Macedonia', 'ms': 'Malaysia', 'ml': 'India', 'mn': 'Mongolia', 'ne': 'Nepal',
                    'no': 'Norway', 'fa': 'Iran', 'pl': 'Poland', 'pt': 'Portugal', 'pa': 'India',
                    'ro': 'Romania', 'ru': 'Russia', 'sr': 'Serbia', 'sk': 'Slovakia', 'sl': 'Slovenia',
                    'es': 'Spain', 'sw': 'Kenya', 'sv': 'Sweden', 'ta': 'India', 'te': 'India',
                    'th': 'Thailand', 'tr': 'Turkey', 'uk': 'Ukraine', 'ur': 'Pakistan', 'uz': 'Uzbekistan',
                    'vi': 'Vietnam', 'xh': 'South Africa', 'zu': 'South Africa'
                }.get(detected_lang, "")
            # 1) Extract emails/phone/facebook from visible text
            text = " ".join(soup.stripped_strings)
            
            # Extract emails using all patterns
            page_emails = set()
            for pattern in email_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    normalized = normalize_obfuscated_email(match)
                    if normalized and re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}$', normalized):
                        page_emails.add(normalized)
            
            # Extract Facebook profiles from text
            facebook_matches = re.findall(facebook_pattern, text, re.IGNORECASE)
            for match in facebook_matches:
                if not match.startswith('http'):
                    match = 'https://' + match
                page_facebook_profiles.add(match)
            
            # Extract phone numbers
            phone_clean = ""
            phones = re.findall(phone_pattern, text)
            for p in phones:
                phone_clean = clean_phone(p)
                if phone_clean:
                    break
            # 2) Extract from mailto: links (first valid)
            mailto_links = [
                a["href"].replace("mailto:", "").strip()
                for a in soup.find_all("a", href=True)
                if "mailto:" in a["href"] and re.search(email_patterns[0], a["href"])
            ]
            for m in mailto_links:
                page_emails.add(clean_email(m))
            tel_links = [
                a["href"].strip()
                for a in soup.find_all("a", href=True)
                if a["href"].lower().startswith("tel:")
            ]
            tel_clean = ""
            for t in tel_links:
                tel_clean = clean_phone(t)
                if tel_clean:
                    break
            # If page has signals, include them in aggregate but keep crawling
            # 3) Discover further links on this page (not just contact-like)
            all_links = set()
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                if href.startswith("mailto:"):
                    continue
                abs_url = urljoin(url, href)
                try:
                    base_host = urlparse(base_url).netloc
                    abs_parsed = urlparse(abs_url)
                    if abs_parsed.scheme in ("http", "https") and abs_parsed.netloc.endswith(base_host):
                        if abs_url not in visited:
                            all_links.add(abs_url)
                except Exception:
                    continue
            # Always prioritize cv/contact links
            prioritized_links = list(cv_contact_links) + [l for l in all_links if l not in cv_contact_links]
            return {
                "all_links": prioritized_links, 
                "emails": list(page_emails), 
                "facebook": list(page_facebook_profiles),
                "phone": tel_clean or phone_clean or None
            }
        except Exception:
            return None

    # Main scraping logic with thread pool
    # Optionally fetch homepage once to get LLM suggestions and seed country/facebook early
    try:
        resp0 = session_get(base_url, timeout=10)
        if resp0.status_code == 200:
            soup0 = BeautifulSoup(resp0.text, 'html.parser')
            # Seed facebook links
            homepage_facebook_profiles = set()
            for a in soup0.find_all("a", href=True):
                href = a["href"].strip()
                if href.startswith("http") and "facebook.com" in href:
                    if facebook_url is None:
                        facebook_url = href
                    homepage_facebook_profiles.add(href)
            # Detect language/country
            if country == "":
                detected_lang = detect_website_language(resp0.text)
                country = {
                    'af': 'South Africa', 'sq': 'Albania', 'am': 'Ethiopia', 'ar': 'Saudi Arabia',
                    'hy': 'Armenia', 'az': 'Azerbaijan', 'eu': 'Spain', 'bn': 'Bangladesh',
                    'bs': 'Bosnia and Herzegovina', 'bg': 'Bulgaria', 'ca': 'Spain', 'zh-cn': 'China',
                    'zh-tw': 'Taiwan', 'hr': 'Croatia', 'cs': 'Czech Republic', 'da': 'Denmark',
                    'nl': 'Netherlands', 'en': 'United States', 'et': 'Estonia', 'fi': 'Finland',
                    'fr': 'France', 'ka': 'Georgia', 'de': 'Germany', 'el': 'Greece', 'gu': 'India',
                    'he': 'Israel', 'hi': 'India', 'hu': 'Hungary', 'is': 'Iceland', 'id': 'Indonesia',
                    'it': 'Italy', 'ja': 'Japan', 'kn': 'India', 'kk': 'Kazakhstan', 'km': 'Cambodia',
                    'ko': 'South Korea', 'ky': 'Kyrgyzstan', 'lo': 'Laos', 'lv': 'Latvia', 'lt': 'Lithuania',
                    'mk': 'North Macedonia', 'ms': 'Malaysia', 'ml': 'India', 'mn': 'Mongolia', 'ne': 'Nepal',
                    'no': 'Norway', 'fa': 'Iran', 'pl': 'Poland', 'pt': 'Portugal', 'pa': 'India',
                    'ro': 'Romania', 'ru': 'Russia', 'sr': 'Serbia', 'sk': 'Slovakia', 'sl': 'Slovenia',
                    'es': 'Spain', 'sw': 'Kenya', 'sv': 'Sweden', 'ta': 'India', 'te': 'India',
                    'th': 'Thailand', 'tr': 'Turkey', 'uk': 'Ukraine', 'ur': 'Pakistan', 'uz': 'Uzbekistan',
                    'vi': 'Vietnam', 'xh': 'South Africa', 'zu': 'South Africa'
                }.get(detected_lang, "")
            # Use LLM to suggest contact pages
            try:
                llm_provider = globals().get('llm_provider_var').get().lower() if globals().get('llm_provider_var') else 'none'
            except Exception:
                llm_provider = 'none'
            llm_links = _llm_suggest_links_from_home(resp0.text, base_url, llm_provider)
            for u in llm_links:
                if u not in visited:
                    visited.add(u)
                    queue.append(u)
            # Also add any obvious links from homepage quickly
            for a in soup0.find_all("a", href=True):
                href = a["href"].strip()
                text_lower = (a.get_text(strip=True) or "").lower()
                href_lower = href.lower()
                if ("cv" in href_lower or "contact" in href_lower or any(k in text_lower for k in ("contact","about","impressum","imprint"))):
                    abs_url = urljoin(base_url, href)
                    try:
                        base_host = urlparse(base_url).netloc
                        abs_parsed = urlparse(abs_url)
                        if abs_parsed.scheme in ("http", "https") and abs_parsed.netloc.endswith(base_host):
                            if abs_url not in visited:
                                visited.add(abs_url)
                                queue.append(abs_url)
                    except Exception:
                        continue
            # Collect emails and Facebook profiles on homepage too
            text0 = " ".join(soup0.stripped_strings)
            
            # Extract emails using all patterns
            for pattern in email_patterns:
                matches = re.findall(pattern, text0, re.IGNORECASE)
                for match in matches:
                    normalized = normalize_obfuscated_email(match)
                    if normalized and re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}$', normalized):
                        found_emails.add(normalized)
            
            # Extract Facebook profiles from homepage text
            facebook_matches = re.findall(facebook_pattern, text0, re.IGNORECASE)
            for match in facebook_matches:
                if not match.startswith('http'):
                    match = 'https://' + match
                found_facebook_profiles.add(match)
            
            # Add homepage Facebook profiles
            found_facebook_profiles.update(homepage_facebook_profiles)
            
            for a in soup0.find_all("a", href=True):
                if a["href"].lower().startswith("mailto:"):
                    found_emails.add(clean_email(a["href"].replace("mailto:", "").strip()))
            if first_phone_global is None:
                phones0 = re.findall(phone_pattern, text0)
                for p in phones0:
                    c = clean_phone(p)
                    if c:
                        first_phone_global = c
                        break
    except Exception:
        pass

    # Build a shared session for connection reuse
    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(pool_connections=threads_var.get(), pool_maxsize=threads_var.get())
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    # Wrap get to use session
    def session_get(u, timeout=10):
        return session.get(u, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(2, threads_var.get())) as executor:
        to_check = list(queue)
        checked = set()
        pages_seen = 0
        max_pages = max(10, maxpages_var.get())
        while to_check and pages_seen < max_pages:
            if hasattr(globals(), 'stop_event') and globals()['stop_event'].is_set():
                break
            # Submit all URLs in to_check
            futures = {executor.submit(fetch_and_parse, url): url for url in to_check}
            to_check = []
            for future in concurrent.futures.as_completed(futures):
                if hasattr(globals(), 'stop_event') and globals()['stop_event'].is_set():
                    break
                result = future.result()
                pages_seen += 1
                if not result:
                    continue
                if isinstance(result, dict):
                    # Aggregate emails, Facebook profiles, and phone
                    for e in result.get("emails", []):
                        if e and re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}$', e):
                            found_emails.add(clean_email(e))
                    for f in result.get("facebook", []):
                        if f:
                            found_facebook_profiles.add(f)
                    if first_phone_global is None and result.get("phone"):
                        first_phone_global = result.get("phone")
                    for link in result.get("all_links", []):
                        if link not in checked and link not in visited:
                            visited.add(link)
                            to_check.append(link)
                checked.add(futures[future])
    # Prepare outputs based on selected options
    result = [base_url]  # Start with website URL
    
    # Add first email if emails are requested
    if scrape_options.get('emails', False):
        first_email = next(iter(found_emails)) if found_emails else None
        result.append(first_email)
    else:
        result.append(None)
    
    # Add phone number if phones are requested
    if scrape_options.get('phone', False):
        result.append(first_phone_global)
    else:
        result.append(None)
    
    # Always add country and facebook (for reference)
    result.extend([country, facebook_url])
    
    # Add count if requested
    if scrape_options.get('count', False):
        if scrape_options.get('emails', False):
            result.append(len(found_emails))
        elif scrape_options.get('facebook', False):
            result.append(len(found_facebook_profiles))
        elif scrape_options.get('phone', False):
            result.append(1 if first_phone_global else 0)
        else:
            result.append(0)
    else:
        result.append(None)
    
    # Add all items if requested
    if scrape_options.get('all', False):
        all_items = []
        if scrape_options.get('emails', False) and found_emails:
            all_items.extend(sorted(found_emails))
        if scrape_options.get('facebook', False) and found_facebook_profiles:
            all_items.extend(sorted(found_facebook_profiles))
        if scrape_options.get('phone', False) and first_phone_global:
            all_items.append(first_phone_global)
        
        result.append("; ".join(all_items) if all_items else None)
    else:
        result.append(None)
    
    return tuple(result)

def update_progress(done, total):
    """Updates progress on the UI."""
    percent = round((done/total)*100, 2) if total else 0
    progressbar['value'] = percent
    progress_var.set(f"Processed: {done}/{total} ({percent}%)")
    _log(f"Processed {done}/{total}")
    root.update_idletasks()

def process_file():
    """Reads the Excel file, extracts URLs, scrapes emails in parallel, and saves results."""
    global file_path
    if not file_path:
        messagebox.showerror("Error", "Please select an Excel file first!")
        return

    try:
        start_row = int(start_entry.get())
        end_row = int(end_entry.get())
        num_processes = int(process_var.get())  # Get selected number of processes
    except ValueError:
        messagebox.showerror("Error", "Invalid input! Please enter valid row numbers.")
        return

    column_letter = column_var.get()
    if not column_letter:
        messagebox.showerror("Error", "Please select a column for URLs!")
        return
    
    # Get scraping options from checkboxes
    scrape_options = {
        'emails': scrape_emails_var.get(),
        'facebook': scrape_facebook_var.get(),
        'phone': scrape_phone_var.get(),
        'count': scrape_count_var.get(),
        'all': scrape_all_var.get()
    }
    
    # Ensure at least one option is selected
    if not any(scrape_options.values()):
        messagebox.showerror("Error", "Please select at least one scraping option!")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    max_row = sheet.max_row

    if start_row < 1 or end_row > max_row or start_row > end_row:
        messagebox.showerror("Error", f"Invalid row range! The file has {max_row} rows.")
        return

    col_idx = string.ascii_uppercase.index(column_letter) + 1
    # Fixed regex: removed null byte
    URL_REGEX = re.compile(r"^(https?|ftp)://[^\s/$.?#].*$")
    def _normalize_url(u: str) -> str:
        u = u.strip()
        # Add https:// if missing scheme
        if not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", u):
            u = "https://" + u
        return u

    urls = []
    url_rows = []  # Track which row each URL came from
    for row in range(start_row, end_row + 1):
        cell_val = sheet.cell(row=row, column=col_idx).value
        if not cell_val:
            urls.append(None)
            url_rows.append(row)
            continue
        candidate = _normalize_url(str(cell_val))
        if URL_REGEX.match(candidate):
            urls.append(candidate)
        else:
            _log(f"Skipping invalid URL at row {row}: {cell_val}")
            urls.append(None)
        url_rows.append(row)
    _log(f"Queued {len(urls)} URLs from row {start_row} to {end_row}")
    total_urls = len(urls)
    completed_urls = 0

    # Update UI
    _set_running_state(True)
    selected_options = [k for k, v in scrape_options.items() if v]
    mode_text = ", ".join([opt.title() for opt in selected_options])
    status_label.config(text=f"Scraping {total_urls} websites for {mode_text} using {num_processes} processes...")
    progressbar['value'] = 0
    progress_var.set("Progress: 0%")
    _log(f"Starting scrape of {total_urls} URLs for {mode_text} with {num_processes} processes")
    root.update()

    def scrape_worker():
        global results_buffer
        results = [None] * len(urls)
        results_buffer = [None] * len(urls)
        # Enable controls during run
        try:
            stop_btn.configure(state="normal")
            download_btn.configure(state="normal")
            resume_btn.configure(state="disabled")
        except Exception:
            pass
        for i, url in enumerate(urls):
            # Pause support
            while pause_event.is_set():
                status_label.config(text="Paused. Click Resume to continue...")
                root.update()
                time.sleep(0.1)
            if url is None:
                results[i] = ("", "", "", "", "", 0, "")
                results_buffer[i] = results[i]
                root.after(0, lambda i=i: _log(f"Row {url_rows[i]}: Invalid or empty link, skipping."))
                root.after(0, lambda i=i: update_progress(i+1, total_urls))
                continue
            root.after(0, lambda i=i, url=url: _log(f"Processing {i+1}/{total_urls}: {url}"))
            result = scrape_email(url, scrape_options)
            results[i] = result
            results_buffer[i] = result
            root.after(0, lambda i=i: update_progress(i+1, total_urls))
        # Save results
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
        save_filename = f"{start_row}_{end_row}_{timestamp}.xlsx"
        save_path = os.path.join(os.path.dirname(file_path), save_filename)
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # Set headers based on selected options
        headers = ["Website"]
        if scrape_options.get('emails', False):
            headers.append("Email")
        if scrape_options.get('phone', False):
            headers.append("Phone")
        headers.extend(["Country", "Facebook"])  # Always include these for reference
        if scrape_options.get('count', False):
            headers.append("Count")
        if scrape_options.get('all', False):
            headers.append("All Items")
        
        new_sheet.append(headers)
        total_found = 0
        for row in results:
            # Normalize row to match header length
            expected_length = len(headers)
            if len(row) < expected_length:
                row = tuple(list(row) + [None]*(expected_length-len(row)))
            elif len(row) > expected_length:
                row = row[:expected_length]
            
            # Count items for total
            try:
                if scrape_options.get('count', False):
                    count_index = headers.index("Count")
                    total_found += int(row[count_index] or 0)
            except Exception:
                pass
            new_sheet.append(row)
        new_wb.save(save_path)
        global last_save_path
        last_save_path = save_path
        # Update success message based on selected options
        selected_options = [k for k, v in scrape_options.items() if v]
        mode_text = ", ".join([opt.title() for opt in selected_options])
        success_msg = f"Scraping complete! Results saved to {save_path}\nTotal items found: {total_found}\nScraped: {mode_text}"
        
        root.after(0, lambda: messagebox.showinfo("Success", success_msg))
        root.after(0, lambda: _auto_open_result(save_path))
        root.after(0, lambda: status_label.config(text="Scraping completed!"))
        # Update progress message based on selected options
        progress_msg = f"Done! Total items found: {total_found}"
        
        root.after(0, lambda: progress_var.set(progress_msg))
        root.after(0, lambda: _log(f"Saved results to: {save_path}"))
        root.after(0, lambda: _set_running_state(False))
        try:
            stop_btn.configure(state="disabled")
            resume_btn.configure(state="disabled")
        except Exception:
            pass
    threading.Thread(target=scrape_worker, daemon=True).start()

def select_file():
    """Opens file dialog to select an Excel file and populate column options."""
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        file_label.config(text=f"Selected: {os.path.basename(file_path)}")
        _log(f"Selected file: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        max_col = sheet.max_column
        # ttk.OptionMenu stores the underlying Menu at ['menu'] as well
        column_menu["menu"].delete(0, "end")
        column_options = list(string.ascii_uppercase[:max_col])

        for col in column_options:
            column_menu["menu"].add_command(label=col, command=tk._setit(column_var, col))
        column_var.set(column_options[0])
        _log(f"Detected columns: {', '.join(column_options)}")

# GUI Setup
root = tk.Tk()
root.title("Professional Email Scraper")
root.geometry("1000x800")
root.configure(bg='#f5f5f5')

# Configure modern light theme with dark buttons
try:
    style = ttk.Style()
    
    # Set light theme
    style.theme_use('clam')
    
    # Configure light color scheme
    light_bg = '#f5f5f5'
    light_fg = '#333333'
    light_select = '#e0e0e0'
    light_entry = '#ffffff'
    accent_blue = '#0078d4'
    accent_green = '#107c10'
    accent_red = '#d13438'
    accent_purple = '#8764b8'
    accent_orange = '#ff8c00'
    
    # Configure root window
    root.configure(bg=light_bg)
    
    # Configure general styles
    style.configure('TLabel', background=light_bg, foreground=light_fg, font=('Segoe UI', 9))
    style.configure('TFrame', background=light_bg)
    style.configure('TLabelFrame', background=light_bg, foreground=light_fg, borderwidth=1, relief='solid')
    style.configure('TLabelFrame.Label', background=light_bg, foreground=light_fg, font=('Segoe UI', 10, 'bold'))
    
    # Configure entry widgets
    style.configure('TEntry', fieldbackground=light_entry, foreground=light_fg, borderwidth=1, insertcolor=light_fg)
    style.map('TEntry', 
              fieldbackground=[('focus', light_select)],
              bordercolor=[('focus', accent_blue)])
    
    # Configure spinbox
    style.configure('TSpinbox', fieldbackground=light_entry, foreground=light_fg, borderwidth=1, insertcolor=light_fg)
    style.map('TSpinbox', 
              fieldbackground=[('focus', light_select)],
              bordercolor=[('focus', accent_blue)])
    
    # Configure option menu
    style.configure('TMenubutton', background=light_entry, foreground=light_fg, borderwidth=1, arrowcolor=light_fg)
    style.map('TMenubutton', 
              background=[('active', light_select)],
              bordercolor=[('focus', accent_blue)])
    
    # Configure checkbuttons
    style.configure('TCheckbutton', background=light_bg, foreground=light_fg, focuscolor='none', font=('Segoe UI', 9))
    style.map('TCheckbutton', 
              background=[('active', light_bg)],
              foreground=[('active', light_fg)])
    
    # Configure progressbar
    style.configure('TProgressbar', background=accent_blue, troughcolor=light_select, borderwidth=0, lightcolor=accent_blue, darkcolor=accent_blue)
    
    # Configure separator
    style.configure('TSeparator', background=light_select)
    
    # Professional button styles
    style.configure("Start.TButton", 
                   foreground="#ffffff", 
                   background=accent_green,
                   font=('Segoe UI', 10, 'bold'),
                   borderwidth=0,
                   focuscolor='none')
    style.map("Start.TButton",
        foreground=[('active', '#ffffff'), ('disabled', '#888888')],
        background=[('active', '#0d6e0d'), ('disabled', '#555555')]
    )
    
    style.configure("Stop.TButton", 
                   foreground="#ffffff", 
                   background=accent_red,
                   font=('Segoe UI', 10, 'bold'),
                   borderwidth=0,
                   focuscolor='none')
    style.map("Stop.TButton",
        foreground=[('active', '#ffffff'), ('disabled', '#888888')],
        background=[('active', '#b52d31'), ('disabled', '#555555')]
    )
    
    style.configure("Resume.TButton", 
                   foreground="#ffffff", 
                   background=accent_blue,
                   font=('Segoe UI', 10, 'bold'),
                   borderwidth=0,
                   focuscolor='none')
    style.map("Resume.TButton",
        foreground=[('active', '#ffffff'), ('disabled', '#888888')],
        background=[('active', '#106ebe'), ('disabled', '#555555')]
    )
    
    style.configure("Download.TButton", 
                   foreground="#ffffff", 
                   background=accent_purple,
                   font=('Segoe UI', 10, 'bold'),
                   borderwidth=0,
                   focuscolor='none')
    style.map("Download.TButton",
        foreground=[('active', '#ffffff'), ('disabled', '#888888')],
        background=[('active', '#7a5ba8'), ('disabled', '#555555')]
    )
    
    style.configure("File.TButton", 
                   foreground="#ffffff", 
                   background=accent_orange,
                   font=('Segoe UI', 9, 'bold'),
                   borderwidth=0,
                   focuscolor='none')
    style.map("File.TButton",
        foreground=[('active', '#ffffff'), ('disabled', '#888888')],
        background=[('active', '#e67e00'), ('disabled', '#555555')]
    )
    
except Exception as e:
    print(f"Style configuration error: {e}")

file_path = ""
last_save_path = ""


# --- Main Frame ---
main_frame = ttk.Frame(root, padding=(12, 10, 12, 12))
main_frame.pack(fill="both", expand=True)

# Header with compact styling
header = ttk.Frame(main_frame)
header.pack(fill="x", pady=(0, 10))

# Main title with icon-like styling
title_frame = ttk.Frame(header)
title_frame.pack(fill="x", pady=(0, 2))
title_lbl = ttk.Label(title_frame, text="📧 Professional Email Scraper", font=("Segoe UI", 10, "bold"))
title_lbl.pack(anchor="w")



# File Selection Section with compact styling
file_section = ttk.LabelFrame(main_frame, text="📁 File Selection", padding=(10, 6))
file_section.pack(fill="x", pady=(0, 8))

file_row = ttk.Frame(file_section)
file_row.pack(fill="x")

ttk.Label(file_row, text="Excel File:", font=("Segoe UI", 9, "bold")).pack(side="left")
choose_btn = ttk.Button(file_row, text="📂 Choose File", command=select_file, style="File.TButton")
choose_btn.pack(side="left", padx=(8, 12))

file_label = ttk.Label(file_row, text="No file selected", foreground="#666666", font=("Segoe UI", 8))
file_label.pack(side="left")

# Options Section with compact styling
opts = ttk.LabelFrame(main_frame, text="⚙️ Configuration Options", padding=(10, 6))
opts.pack(fill="x", pady=(0, 8))

# Create a grid with compact spacing
grid = ttk.Frame(opts)
grid.pack(fill="x", padx=3, pady=3)

# Configure grid weights for better layout
for i in range(6):
    grid.columnconfigure(i, weight=1)

# Theme and UI Scale
ttk.Label(grid, text="Theme:").grid(row=0, column=2, sticky="w", padx=(12, 6), pady=4)
theme_var = tk.StringVar(root)
try:
    style = ttk.Style()
    themes = [t for t in ("vista", "clam", "default") if t in style.theme_names()]
except Exception:
    themes = ["default"]
theme_var.set(themes[0] if themes else "default")
theme_menu = ttk.OptionMenu(grid, theme_var, theme_var.get(), *themes)
theme_menu.grid(row=0, column=3, sticky="w", pady=4)

def _apply_theme(*_):
    try:
        ttk.Style().theme_use(theme_var.get())
    except Exception:
        pass
theme_var.trace_add('write', _apply_theme)

ttk.Label(grid, text="UI Scale:").grid(row=1, column=2, sticky="w", padx=(8, 4), pady=4)
scale_var = tk.DoubleVar(root, value=1.0)
def _apply_scale(*_):
    try:
        root.tk.call('tk', 'scaling', scale_var.get())
    except Exception:
        pass
scale_spin = ttk.Spinbox(grid, from_=0.8, to=1.6, increment=0.1, textvariable=scale_var, width=6, command=lambda: _apply_scale())
scale_spin.grid(row=1, column=3, sticky="w", pady=4)
scale_var.trace_add('write', _apply_scale)

ttk.Label(grid, text="URL Column:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
column_var = tk.StringVar(root)
column_menu = ttk.OptionMenu(grid, column_var, "")
column_menu.grid(row=0, column=1, sticky="w", pady=4)

# Scraping options section
scrape_options_label = ttk.Label(grid, text="🎯 Scraping Options:", font=("Segoe UI", 9, "bold"))
scrape_options_label.grid(row=0, column=4, sticky="w", padx=(15, 6), pady=4)

# Create a frame for checkboxes with compact styling
checkbox_frame = ttk.LabelFrame(grid, text="Select Data Types", padding=(6, 4))
checkbox_frame.grid(row=0, column=5, sticky="ew", padx=(0, 0), pady=2, columnspan=2)

# Checkbox variables
scrape_emails_var = tk.BooleanVar(root, value=True)
scrape_facebook_var = tk.BooleanVar(root, value=False)
scrape_phone_var = tk.BooleanVar(root, value=False)
scrape_count_var = tk.BooleanVar(root, value=True)
scrape_all_var = tk.BooleanVar(root, value=True)

# Create checkboxes with icons and compact styling
email_cb = ttk.Checkbutton(checkbox_frame, text="📧 Emails", variable=scrape_emails_var)
email_cb.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=1)

facebook_cb = ttk.Checkbutton(checkbox_frame, text="📘 Facebook", variable=scrape_facebook_var)
facebook_cb.grid(row=0, column=1, sticky="w", padx=(0, 10), pady=1)

phone_cb = ttk.Checkbutton(checkbox_frame, text="📞 Phone", variable=scrape_phone_var)
phone_cb.grid(row=0, column=2, sticky="w", padx=(0, 10), pady=1)

count_cb = ttk.Checkbutton(checkbox_frame, text="🔢 Count", variable=scrape_count_var)
count_cb.grid(row=1, column=0, sticky="w", padx=(0, 10), pady=1)

all_cb = ttk.Checkbutton(checkbox_frame, text="📋 All Items", variable=scrape_all_var)
all_cb.grid(row=1, column=1, sticky="w", padx=(0, 10), pady=1)

ttk.Label(grid, text="Start Row:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=4)
start_entry = ttk.Entry(grid, width=12)
start_entry.grid(row=1, column=1, sticky="w", pady=4)
start_entry.insert(0, "1")

ttk.Label(grid, text="End Row:").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=4)
end_entry = ttk.Entry(grid, width=12)
end_entry.grid(row=2, column=1, sticky="w", pady=4)

max_processes = cpu_count() * 2
ttk.Label(grid, text=f"Processes (Max: {cpu_count()}):").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=4)
process_var = tk.StringVar(root)
process_options = [str(2**i) for i in range(1, max_processes.bit_length()) if 2**i <= max_processes]
process_var.set(str(max(1, cpu_count() // 2)))
process_menu = ttk.OptionMenu(grid, process_var, process_var.get(), *process_options)
process_menu.grid(row=3, column=1, sticky="w", pady=4)

# Speed Controls
ttk.Label(grid, text="Threads/Site:").grid(row=2, column=2, sticky="w", padx=(12, 6), pady=4)
threads_var = tk.IntVar(root, value=8)
threads_spin = ttk.Spinbox(grid, from_=2, to=32, increment=1, textvariable=threads_var, width=6)
threads_spin.grid(row=2, column=3, sticky="w", pady=4)

ttk.Label(grid, text="Max Pages/Site:").grid(row=3, column=2, sticky="w", padx=(12, 6), pady=4)
maxpages_var = tk.IntVar(root, value=60)
maxpages_spin = ttk.Spinbox(grid, from_=10, to=500, increment=10, textvariable=maxpages_var, width=6)
maxpages_spin.grid(row=3, column=3, sticky="w", pady=4)

# Separator
ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=(4, 8))

# AI Options Group with compact styling
ai_frame = ttk.LabelFrame(main_frame, text="🤖 AI Options", padding=(10, 6))
ai_frame.pack(fill="x", pady=(0, 8))
ai_grid = ttk.Frame(ai_frame)
ai_grid.pack(fill="x", padx=5, pady=4)

ttk.Label(ai_grid, text="LLM Boost:").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
llm_provider_var = tk.StringVar(root)
llm_choices = ["None", "OpenAI", "Gemini"]
llm_provider_var.set("None")
llm_menu = ttk.OptionMenu(ai_grid, llm_provider_var, llm_provider_var.get(), *llm_choices)
llm_menu.grid(row=0, column=1, sticky="w", pady=4)

ttk.Label(ai_grid, text="OpenAI Key:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=4)
openai_key_var = tk.StringVar(root)
openai_key_entry = ttk.Entry(ai_grid, width=40, textvariable=openai_key_var, show="*")
openai_key_entry.grid(row=1, column=1, sticky="w", pady=4)

ttk.Label(ai_grid, text="Gemini Key:").grid(row=2, column=0, sticky="w", padx=(0, 6), pady=4)
gemini_key_var = tk.StringVar(root)
gemini_key_entry = ttk.Entry(ai_grid, width=40, textvariable=gemini_key_var, show="*")
gemini_key_entry.grid(row=2, column=1, sticky="w", pady=4)

def _refresh_llm_status():
    have_openai = bool((openai_key_var.get() or OPENAI_API_KEY))
    have_gemini = bool((gemini_key_var.get() or GEMINI_API_KEY))
    parts = []
    if have_openai:
        parts.append("OpenAI")
    if have_gemini:
        parts.append("Gemini")
    txt = ("Keys: " + ", ".join(parts)) if parts else "Keys: none"
    llm_status.configure(text=txt)

llm_status = ttk.Label(ai_grid, text="Keys: none")
llm_status.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0,4))
openai_key_var.trace_add('write', lambda *args: _refresh_llm_status())
gemini_key_var.trace_add('write', lambda *args: _refresh_llm_status())

# Separator
ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=(4, 8))

# Actions Section with compact styling
actions_section = ttk.LabelFrame(main_frame, text="🚀 Actions", padding=(10, 6))
actions_section.pack(fill="x", pady=(0, 8))

actions = ttk.Frame(actions_section)
actions.pack(fill="x")

# Main action buttons with compact spacing
start_btn = ttk.Button(actions, text="▶️ Start Scraping", command=lambda: process_file(), style="Start.TButton")
start_btn.pack(side="left", padx=(0, 8))
# --- New Buttons ---
def on_stop():
    pause_event.set()
    try:
        stop_btn.configure(state="disabled")
        resume_btn.configure(state="normal")
        status_label.config(text="Paused. Click Resume to continue...")
    except Exception:
        pass

def on_resume():
    pause_event.clear()
    try:
        stop_btn.configure(state="normal")
        resume_btn.configure(state="disabled")
        status_label.config(text="Resumed.")
    except Exception:
        pass

def _write_results_xlsx(path, rows, scrape_options=None):
    if scrape_options is None:
        scrape_options = {'emails': True, 'facebook': False, 'phone': False, 'count': True, 'all': True}
    
    wb = openpyxl.Workbook()
    sh = wb.active
    
    # Set headers based on selected options
    headers = ["Website"]
    if scrape_options.get('emails', False):
        headers.append("Email")
    if scrape_options.get('phone', False):
        headers.append("Phone")
    headers.extend(["Country", "Facebook"])  # Always include these for reference
    if scrape_options.get('count', False):
        headers.append("Count")
    if scrape_options.get('all', False):
        headers.append("All Items")
    
    sh.append(headers)
    
    for r in rows:
        if not r:
            continue
        # Normalize row to match header length
        expected_length = len(headers)
        if len(r) < expected_length:
            r = tuple(list(r) + [None]*(expected_length-len(r)))
        elif len(r) > expected_length:
            r = r[:expected_length]
        sh.append(r)
    wb.save(path)

def on_download_now():
    try:
        if not results_buffer:
            messagebox.showinfo("Download", "No results to download yet.")
            return
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
        dl_name = f"partial_{timestamp}.xlsx"
        dl_path = os.path.join(os.path.dirname(file_path) if file_path else os.getcwd(), dl_name)
        rows = [r for r in results_buffer if r]
        # Get current scraping options
        current_options = {
            'emails': scrape_emails_var.get(),
            'facebook': scrape_facebook_var.get(),
            'phone': scrape_phone_var.get(),
            'count': scrape_count_var.get(),
            'all': scrape_all_var.get()
        }
        
        _write_results_xlsx(dl_path, rows, current_options)
        _auto_open_result(dl_path)
        _log(f"Partial results exported to: {dl_path}")
        status_bar.configure(text=f"Exported partial results: {dl_name}")
    except Exception as e:
        messagebox.showerror("Error", f"Download failed: {e}")

stop_btn = ttk.Button(actions, text="⏸️ Stop", state="disabled", command=on_stop, style="Stop.TButton")
stop_btn.pack(side="left", padx=(0, 8))
resume_btn = ttk.Button(actions, text="▶️ Resume", state="disabled", command=on_resume, style="Resume.TButton")
resume_btn.pack(side="left", padx=(0, 8))
download_btn = ttk.Button(actions, text="💾 Download Now", state="disabled", command=on_download_now, style="Download.TButton")
download_btn.pack(side="left", padx=(0, 8))

# Progress Section with compact styling
progress_section = ttk.LabelFrame(main_frame, text="📊 Progress", padding=(10, 6))
progress_section.pack(fill="x", pady=(0, 8))

progress_frame = ttk.Frame(progress_section)
progress_frame.pack(fill="x", pady=(0, 4))

status_label = ttk.Label(progress_frame, text="⏳ Waiting for input...", font=("Segoe UI", 9, "bold"))
status_label.pack(side="left")

# Compact progress bar
progressbar = ttk.Progressbar(progress_section, orient="horizontal", mode="determinate", maximum=100)
progressbar.pack(fill="x", pady=(0, 4))

progress_var = tk.StringVar(root)
progress_var.set("Progress: 0%")
progress_label = ttk.Label(progress_section, textvariable=progress_var, font=("Segoe UI", 8))
progress_label.pack(anchor="w")

# Log Panel with compact styling
log_frame = ttk.LabelFrame(main_frame, text="📝 Activity Log", padding=(8, 6))
log_frame.pack(fill="both", expand=True, pady=(0, 8))

# Configure log text with light theme
log_text = ScrolledText(log_frame, height=6, wrap="word", state="disabled", 
                       bg='#ffffff', fg='#333333', insertbackground='#333333',
                       selectbackground='#e0e0e0', selectforeground='#333333',
                       font=('Consolas', 8))
log_text.pack(fill="both", expand=True, padx=4, pady=4)

# Status Bar with light theme styling
status_bar = ttk.Label(root, text="💡 Tip: Provide an LLM key to improve discovery and enable AI-powered contact page detection", 
                      anchor="w", font=("Segoe UI", 8), foreground="#666666")
status_bar.pack(side="bottom", fill="x", padx=8, pady=3)

# --- Control Variables for Stop/Resume/Download ---
import threading
scrape_thread = None
stop_event = threading.Event()
pause_event = threading.Event()
results_buffer = []  # Holds current results for download

# Enhanced menubar with light theme
menubar = tk.Menu(root, bg='#f5f5f5', fg='#333333', activebackground='#e0e0e0', activeforeground='#333333')
help_menu = tk.Menu(menubar, tearoff=0, bg='#f5f5f5', fg='#333333', activebackground='#e0e0e0', activeforeground='#333333')

def _about():
    about_text = """📧 Professional Email Scraper v2.0

✨ Features:
• Extract emails, Facebook profiles, and phone numbers
• Handle obfuscated email formats
• AI-powered contact page discovery
• Multiple data type selection
• Professional dark theme UI
• Real-time progress tracking

🔧 Built with Python & Tkinter"""
    messagebox.showinfo("About Professional Email Scraper", about_text)

help_menu.add_command(label="📖 About", command=_about)
menubar.add_cascade(label="ℹ️ Help", menu=help_menu)
root.config(menu=menubar)

# On-close behavior
root.protocol("WM_DELETE_WINDOW", _on_close)

if __name__ == "__main__":
    freeze_support()  # Prevents multiple windows on Windows
    root.mainloop()
